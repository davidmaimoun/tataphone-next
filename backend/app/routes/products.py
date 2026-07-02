import openpyxl
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt, verify_jwt_in_request
from app.services.r2_storage import upload_to_r2

from app.models import product as ProductModel

products_bp = Blueprint('products', __name__)

 

def _sync_meta(product_data):
    """Auto-add brand/category/colors/sizes to meta collections."""
    try:
        from app.db import get_db
        import datetime
        db  = get_db()
        now = datetime.datetime.utcnow()
        entries = {
            'meta_brands':     [product_data.get('brand')],
            'meta_categories': [product_data.get('category')],
            'meta_colors':     product_data.get('colors', []),
            'meta_sizes':      product_data.get('sizes', []),
        }
        for col, names in entries.items():
            for name in (names or []):
                if name and str(name).strip():
                    db[col].update_one(
                        {'name': str(name).strip()},
                        {'$setOnInsert': {'name': str(name).strip(), 'createdAt': now}},
                        upsert=True
                    )
    except Exception as e:
        print(f'[META] sync failed: {e}')

ALLOWED_IMG = {'png', 'jpg', 'jpeg', 'webp', 'gif'}


def _is_admin():
    try:
        verify_jwt_in_request()
        return get_jwt().get('role','user') == 'admin'
    except Exception:
        return False


def _allowed_img(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMG



# ── 2) REMPLACE ta fonction _save_image(...) par : ──
def _save_image(file) -> str:
    """Upload vers Cloudflare R2 (resize Pillow inclus). Retourne l'URL publique."""
    return upload_to_r2(file)
 

# ── GET /api/products ─────────────────────────────────────────────────────────
@products_bp.route('/', methods=['GET'])
def list_products():
    params = request.args.to_dict()
    result = ProductModel.get_all(params)
    return jsonify(result)


# ── POST /api/products/recommended ───────────────────────────────────────────
@products_bp.route('/recommended', methods=['POST'])
def recommended():
    visited = (request.get_json() or {}).get('visitedIds', [])
    products = ProductModel.get_recommended(visited)
    return jsonify({'products': products})


# ── POST /api/products/import ────────────────────────────────────────────────
@products_bp.route('/import', methods=['POST'])
@jwt_required()
def import_excel():
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400

    file = request.files['file']
    wb   = openpyxl.load_workbook(file, data_only=True)
    ws   = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
    imported = 0
    errors   = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)}
            # Map common CRM column aliases
            data.setdefault('name',      data.pop('שם', data.get('product_name', '')))
            data.setdefault('brand',     data.pop('מותג', data.get('manufacturer', '')))
            data.setdefault('category',  data.pop('קטגוריה', data.get('type', 'smartphones')))
            data.setdefault('price',     data.pop('מחיר', data.get('price_ils', 0)))
            data.setdefault('sku',       data.pop('מק"ט', data.get('sku', '')))

            if not data.get('name'):
                continue

            ProductModel.create_product(data)
            imported += 1
        except Exception as e:
            errors.append({'row': row_idx, 'error': str(e)})

    return jsonify({'imported': imported, 'errors': errors})



@products_bp.route('/bestsellers', methods=['GET'])
def bestsellers():
    from app.db import get_db
    limit = min(24, max(1, int(request.args.get('limit', 8))))
    db = get_db()
    # Featured d'abord, puis par ventes décroissantes, puis récents
    cursor = db['products'].find({}).sort([
        ('isFeatured', -1),   # featured en tête
        ('salesCount', -1),   # puis meilleures ventes
        ('createdAt', -1),    # fallback : récents (si tout à 0)
    ]).limit(limit)
    products = [ProductModel.serialize(p) for p in cursor]
    return jsonify({'products': products})
 
 
@products_bp.route('/top-rated', methods=['GET'])
def top_rated():
    from app.db import get_db
    limit = min(24, max(1, int(request.args.get('limit', 8))))
    db = get_db()
    cursor = db['products'].find({}).sort([
        ('isFeatured', -1),   # featured en tête
        ('rating', -1),       # puis meilleure note
        ('salesCount', -1),   # départage par ventes
        ('createdAt', -1),    # fallback : récents
    ]).limit(limit)
    products = [ProductModel.serialize(p) for p in cursor]
    return jsonify({'products': products})
 
 
# ── GET /api/products/:id ─────────────────────────────────────────────────────
@products_bp.route('/<product_id>', methods=['GET'])
def get_product(product_id):
    p = ProductModel.get_by_id(product_id)
    if not p:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(p)


# ── GET /api/products/:id/related ────────────────────────────────────────────
@products_bp.route('/<product_id>/related', methods=['GET'])
def related(product_id):
    products = ProductModel.get_related(product_id)
    return jsonify({'products': products})


# ── POST /api/products ────────────────────────────────────────────────────────
@products_bp.route('/', methods=['POST'])
@jwt_required()
def create_product():
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403

    data   = request.form.to_dict()
    images = []

    # New photos uploaded as files
    for f in request.files.getlist('photos'):
        if f and _allowed_img(f.filename):
            images.append(_save_image(f))

    # Legacy field name support
    for f in request.files.getlist('images'):
        if f and _allowed_img(f.filename):
            images.append(_save_image(f))

    data['images'] = images
    p = ProductModel.create_product(data)
    _sync_meta(p)
    return jsonify(p), 201


# ── DELETE /api/products/:id ── (REMPLACE l'existante)
@products_bp.route('/<product_id>', methods=['DELETE'])
@jwt_required()
def delete_product(product_id):
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403
    # Récupère le produit AVANT suppression pour connaître ses images
    product = ProductModel.get_by_id(product_id)
    ProductModel.delete_product(product_id)
    # Supprime les images du bucket R2
    if product:
        from app.services.r2_storage import delete_from_r2
        for img_url in (product.get('images') or []):
            try:
                delete_from_r2(img_url)
            except Exception as e:
                print(f"[R2] delete image failed: {e}")
    return jsonify({'deleted': True})
 
 
# ── PUT /api/products/:id ── (REMPLACE l'existante)
#  Supprime du bucket les images RETIRÉES lors de l'update.
@products_bp.route('/<product_id>', methods=['PUT'])
@jwt_required()
def update_product(product_id):
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403
 
    import json
    from app.services.r2_storage import delete_from_r2
    data = request.form.to_dict() if request.form else (request.get_json() or {})
 
    # Images existantes AVANT update (pour détecter celles retirées)
    old_product = ProductModel.get_by_id(product_id)
    old_images  = set((old_product or {}).get('images', []))
 
    # Images conservées (envoyées par le front)
    existing = json.loads(data.pop('existingImages', '[]'))
    new_imgs = list(existing)
 
    # Upload nouvelles photos
    for f in request.files.getlist('photos'):
        if f and _allowed_img(f.filename):
            new_imgs.append(_save_image(f))
 
    data['images'] = new_imgs
    p = ProductModel.update_product(product_id, data)
    _sync_meta(p)
 
    # ── Nettoyage : supprimer du bucket les images qui ne sont plus utilisées ──
    kept = set(new_imgs)
    removed = old_images - kept
    for img_url in removed:
        try:
            delete_from_r2(img_url)
        except Exception as e:
            print(f"[R2] delete removed image failed: {e}")
 
    return jsonify(p)

# ── POST /api/products/:id/photos ─────────────────────────────────────────────
@products_bp.route('/<product_id>/photos', methods=['POST'])
@jwt_required()
def upload_photos(product_id):
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403

    urls = []
    for f in request.files.getlist('photos'):
        if f and _allowed_img(f.filename):
            urls.append(_save_image(f))

    if not urls:
        return jsonify({'error': 'No valid images'}), 400

    ProductModel.add_images(product_id, urls)
    p = ProductModel.get_by_id(product_id)
    return jsonify({'images': p['images']})


# ── GET /api/products/meta/brands ─────────────────────────────────────────────
@products_bp.route('/meta/brands', methods=['GET'])
def meta_brands():
    from app.db import get_db
    brands = get_db()['products'].distinct('brand')
    return jsonify({'brands': sorted([b for b in brands if b])})


# ── GET /api/products/meta/colors ─────────────────────────────────────────────
@products_bp.route('/meta/colors', methods=['GET'])
def meta_colors():
    from app.db import get_db
    # Collect all colors from products that have colors array
    pipeline = [
        {'$unwind': '$colors'},
        {'$group': {'_id': '$colors'}},
        {'$sort': {'_id': 1}},
    ]
    result = list(get_db()['products'].aggregate(pipeline))
    colors = [r['_id'] for r in result if r['_id']]
    if not colors:
        colors = ['שחור','לבן','כסף','זהב','כחול','אדום','ירוק','אפור','ורוד','סגול']
    return jsonify({'colors': colors})


# ── POST /api/products/import-json ────────────────────────────────────────────
@products_bp.route('/import-json', methods=['POST'])
@jwt_required()
def import_json():
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403
    body = request.get_json() or {}
    products_data = body.get('products', [])
    overwrite = body.get('overwriteSku', False)
    from app.db import get_db
    col = get_db()['products']
    imported, updated, errors = 0, 0, []
    for i, p in enumerate(products_data):
        try:
            if not p.get('name') or not p.get('price'):
                errors.append({'row': i+1, 'error': 'Missing name or price'}); continue
            sku = (p.get('sku') or '').strip()
            if overwrite and sku:
                existing = col.find_one({'sku': sku})
                if existing:
                    ProductModel.update_product(str(existing['_id']), p)
                    updated += 1; continue
            ProductModel.create_product(p)
            imported += 1
        except Exception as e:
            errors.append({'row': i+1, 'error': str(e)})
    return jsonify({'imported': imported, 'updated': updated, 'errors': errors})
 

# ── POST /api/products/by-ids ─────────────────────────────────────────────────
@products_bp.route('/by-ids', methods=['POST'])
def get_by_ids():
    """Public endpoint — fetch products by list of IDs (for guest wishlist)."""
    ids = (request.get_json() or {}).get('ids', [])
    if not ids:
        return jsonify({'products': []})
    from bson import ObjectId
    results = []
    for pid in ids[:50]:  # limit 50
        try:
            p = ProductModel.get_by_id(pid)
            if p:
                results.append(p)
        except Exception:
            pass
    return jsonify({'products': results})


@products_bp.route('/import-schema', methods=['GET'])
def import_schema():
    from app.db import get_db
    db = get_db()
    fields = [
        {'key':'name',          'label':'שם המוצר',     'required':True,  'type':'text'},
        {'key':'price',         'label':'מחיר',          'required':True,  'type':'number'},
        {'key':'brand',         'label':'מותג',          'required':False, 'type':'text'},
        {'key':'sku',           'label':'מק"ט',          'required':False, 'type':'text'},
        {'key':'category',      'label':'קטגוריה',       'required':False, 'type':'text', 'default':'smartphones'},
        {'key':'description',   'label':'תיאור',         'required':False, 'type':'text'},
        {'key':'originalPrice', 'label':'מחיר מקורי',    'required':False, 'type':'number', 'note':'גבוה מהמחיר → מבצע'},
        {'key':'supplierPrice', 'label':'מחיר ספק',      'required':False, 'type':'number', 'note':'פנימי — לא מוצג ללקוח'},
        {'key':'stock',         'label':'מלאי',          'required':False, 'type':'number'},
        {'key':'rating',        'label':'דירוג',         'required':False, 'type':'number'},
        {'key':'reviewCount',   'label':'מספר ביקורות',  'required':False, 'type':'number'},
        {'key':'isKosher',      'label':'כשר',           'required':False, 'type':'boolean', 'note':'yes/no, כן/לא'},
        {'key':'isAccessory',   'label':'אביזר',         'required':False, 'type':'boolean'},
        {'key':'isNew',         'label':'חדש',           'required':False, 'type':'boolean'},
        {'key':'isTopRated',    'label':'מדורג גבוה',    'required':False, 'type':'boolean'},
        {'key':'tags',          'label':'תגיות',         'required':False, 'type':'list'},
        {'key':'selectedColors','label':'צבעים',         'required':False, 'type':'list'},
        {'key':'selectedSizes', 'label':'מידות / אחסון', 'required':False, 'type':'list'},
        {'key':'specs',         'label':'מפרט טכני',     'required':False, 'type':'object'},
        {'key':'note',          'label':'הערת מנהל',     'required':False, 'type':'text'},
        {'key':'images',        'label':'תמונות (URLs)', 'required':False, 'type':'list'},
        {'key':'details', 'label':'מקטעים (תיאור/אחריות/משלוח)', 'required':False, 'type':'list',
         'note':'מערך של {title, body} — לשוניות מתחת למוצר'},
    ]
    try:
        cats   = sorted([c for c in db['products'].distinct('category') if c])
        brands = sorted([b for b in db['products'].distinct('brand') if b])
        tags   = sorted([t for t in db['products'].distinct('tags') if t])
        total  = db['products'].count_documents({})
    except Exception:
        cats, brands, tags, total = [], [], [], 0
    return jsonify({'fields': fields, 'existingCategories': cats,
                    'existingBrands': brands, 'existingTags': tags, 'totalProducts': total})
 
 
@products_bp.route('/admin/list', methods=['GET'])
@jwt_required()
def admin_list():
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403
    from app.db import get_db
    col = get_db()['products']
    products = [ProductModel.serialize(p, admin=True) for p in col.find().sort('createdAt', -1)]
    return jsonify({'products': products, 'total': len(products)})


@products_bp.route('/export', methods=['GET'])
@jwt_required()
def export_products():
    if not _is_admin():
        return jsonify({'error': 'Admin only'}), 403
 
    from flask import send_file
    from app.db import get_db
    import io, json, csv, datetime
 
    fmt = request.args.get('format', 'json').lower()
    products = list(get_db()['products'].find())
    stamp = datetime.datetime.utcnow().strftime('%Y-%m-%d_%H-%M')
 
    # Nettoyer pour l'export (retirer _id ObjectId, dates → str)
    clean = []
    for p in products:
        p.pop('_id', None)
        for k in ('createdAt', 'updatedAt'):
            if k in p and hasattr(p[k], 'isoformat'):
                p[k] = p[k].isoformat()
        clean.append(p)
 
    if fmt == 'csv':
        # CSV : colonnes principales (variants/options aplaties en JSON string)
        output = io.StringIO()
        cols = ['sku','name','nameEn','brand','category','subCategory','price',
                'originalPrice','supplierPrice','stock','isKosher','isAccessory',
                'description','barcode','supplier','options','variants','tags']
        writer = csv.DictWriter(output, fieldnames=cols, extrasaction='ignore')
        writer.writeheader()
        for p in clean:
            row = dict(p)
            # Sérialiser les champs complexes en JSON string
            for cx in ('options','variants','tags','images','specs','details'):
                if cx in row and not isinstance(row[cx], str):
                    row[cx] = json.dumps(row[cx], ensure_ascii=False)
            writer.writerow(row)
        # BOM UTF-8 pour Excel (affiche bien l'hébreu)
        data = '\ufeff' + output.getvalue()
        return send_file(
            io.BytesIO(data.encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'tataphone_products_{stamp}.csv'
        )
    else:
        # JSON (réimportable)
        data = json.dumps(clean, ensure_ascii=False, indent=2)
        return send_file(
            io.BytesIO(data.encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'tataphone_products_{stamp}.json'
        )
 
